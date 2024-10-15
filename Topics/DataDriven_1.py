# Data Driven Testing using Excel file
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import ActionChains

from openpyxl import load_workbook

from time import sleep

chrome_driver = r"C:\selenium\chromedriver.exe"
service = Service(chrome_driver)
options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
driver = webdriver.Chrome(service=service, options=options)

# url = "https://www.saucedemo.com/v1/"
# driver.get(url)
driver.maximize_window()

# Load Excel Sheet
workbook = load_workbook('test_data.xlsx')

# Selecting Active Sheet
wb = workbook.active
for row in wb.iter_rows(min_row=2, max_row=wb.max_row, values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/")
    sleep(2)
    driver.find_element(By.ID, 'user-name').send_keys(username)
    driver.find_element(By.ID, 'password').send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    sleep(2)
    driver.find_element(By.XPATH, "//button[normalize-space()='Open Menu']").click()
    sleep(2)
    driver.find_element(By.XPATH, "//a[@id='logout_sidebar_link']").click()
    sleep(2)
driver.quit()
